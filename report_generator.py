import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, PageBreak, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
import config
from sqlalchemy.orm import joinedload
from database import (
    SessionLocal, Drill, DrillStep, DrillIssue, DrillReport,
    MonthlyReport, ImprovementWorkOrder, BusinessFunction, BusinessContinuityPlan
)
from logging_system import report_logger, log_system_event


def _setup_chinese_font():
    available_fonts = [f.name for f in fm.fontManager.ttflist]
    chinese_fonts = ["PingFang SC", "Heiti SC", "STHeiti", "SimHei", "Microsoft YaHei", "Arial Unicode MS"]
    for font in chinese_fonts:
        if font in available_fonts:
            plt.rcParams["font.sans-serif"] = [font]
            plt.rcParams["axes.unicode_minus"] = False
            return font
    plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    return "DejaVu Sans"


_setup_chinese_font()


def generate_trend_chart(data: List[Dict], title: str, x_key: str, y_key: str, output_path: str) -> Optional[str]:
    try:
        fig, ax = plt.subplots(figsize=(10, 6))
        
        x_values = [d[x_key] for d in data]
        y_values = [d[y_key] for d in data]
        
        ax.plot(x_values, y_values, marker='o', linewidth=2, markersize=6, color='#2563eb')
        ax.fill_between(range(len(x_values)), y_values, alpha=0.15, color='#2563eb')
        
        ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel(x_key, fontsize=11)
        ax.set_ylabel(y_key, fontsize=11)
        ax.grid(True, alpha=0.3)
        ax.tick_params(axis='x', rotation=45)
        
        for i, v in enumerate(y_values):
            ax.annotate(f'{v:.1f}', (i, v), textcoords="offset points", xytext=(0, 10), ha='center', fontsize=9)
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
        
        report_logger.info(f"Trend chart generated: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to generate trend chart: {str(e)}")
        return None


def generate_comparison_chart(data: Dict, title: str, output_path: str) -> Optional[str]:
    try:
        fig, ax = plt.subplots(figsize=(12, 7))
        
        categories = list(data.keys())
        values = list(data.values())
        
        bars = ax.bar(categories, values, color=['#2563eb', '#16a34a', '#f59e0b', '#dc2626', '#7c3aed', '#0891b2', '#65a30d'])
        
        ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
        ax.set_ylabel('数值', fontsize=11)
        ax.tick_params(axis='x', rotation=30)
        ax.grid(True, alpha=0.3, axis='y')
        
        for bar, value in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2., bar.get_height(),
                    f'{value:.1f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
        
        report_logger.info(f"Comparison chart generated: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to generate comparison chart: {str(e)}")
        return None


def analyze_drill_results(drill_id: int) -> Dict:
    db = SessionLocal()
    try:
        drill = db.query(Drill).get(drill_id)
        if not drill:
            return {}
        
        steps = db.query(DrillStep).filter(DrillStep.drill_id == drill_id).all()
        issues = db.query(DrillIssue).filter(DrillIssue.drill_id == drill_id).all()
        
        total_steps = len(steps)
        completed_steps = sum(1 for s in steps if s.status == "completed")
        overdue_steps = sum(1 for s in steps if s.is_overdue)
        
        recovery_success_rate = (completed_steps / total_steps * 100) if total_steps > 0 else 0
        
        recovery_times = []
        time_deviations = []
        for step in steps:
            if step.actual_start_time and step.actual_end_time:
                actual_duration = (step.actual_end_time - step.actual_start_time).total_seconds() / 60
                recovery_times.append(actual_duration)
                if step.target_duration_minutes:
                    deviation = actual_duration - step.target_duration_minutes
                    time_deviations.append(deviation)
        
        avg_recovery_time = sum(recovery_times) / len(recovery_times) if recovery_times else 0
        avg_time_deviation = sum(time_deviations) / len(time_deviations) if time_deviations else 0
        
        problem_list = []
        for step in steps:
            if step.is_overdue:
                problem_list.append({
                    "type": "step_overdue",
                    "step": step.step_number,
                    "description": step.description,
                    "severity": "high" if step.target_duration_minutes and (step.actual_end_time and (step.actual_end_time - step.actual_start_time).total_seconds() / 60 > step.target_duration_minutes * 2) else "medium"
                })
        
        for issue in issues:
            problem_list.append({
                "type": "issue_found",
                "issue_id": issue.id,
                "description": issue.description,
                "severity": issue.severity
            })
        
        recommendations = []
        if recovery_success_rate < 80:
            recommendations.append("演练成功率低于80%，建议重新评估预案的可行性并增加培训")
        if avg_time_deviation > 10:
            recommendations.append(f"平均时间偏差{avg_time_deviation:.1f}分钟，建议优化恢复流程或增加资源投入")
        if overdue_steps > 0:
            recommendations.append(f"共有{overdue_steps}个步骤超时，建议对相关人员进行专项培训")
        if any(p["severity"] == "critical" for p in problem_list):
            recommendations.append("存在严重问题，建议立即召开复盘会议并制定整改计划")
        
        critical_issues = sum(1 for i in issues if i.severity == "critical")
        
        return {
            "drill_id": drill.id,
            "drill_name": drill.name,
            "drill_type": drill.drill_type,
            "recovery_success_rate": round(recovery_success_rate, 1),
            "avg_recovery_time": round(avg_recovery_time, 1),
            "time_deviation": round(avg_time_deviation, 1),
            "steps_completed": completed_steps,
            "total_steps": total_steps,
            "issues_found": len(issues),
            "critical_issues": critical_issues,
            "problem_list": problem_list,
            "recommendations": recommendations,
            "start_time": drill.actual_start_time.isoformat() if drill.actual_start_time else None,
            "end_time": drill.actual_end_time.isoformat() if drill.actual_end_time else None,
            "target_recovery_time": drill.target_recovery_time
        }
    except Exception as e:
        report_logger.error(f"Failed to analyze drill results: {str(e)}", exc_info=True)
        return {}
    finally:
        db.close()


def save_drill_report(drill_id: int) -> Optional[DrillReport]:
    db = SessionLocal()
    try:
        analysis = analyze_drill_results(drill_id)
        if not analysis:
            return None
        
        report = DrillReport(
            drill_id=drill_id,
            recovery_success_rate=analysis["recovery_success_rate"],
            avg_recovery_time=analysis["avg_recovery_time"],
            time_deviation=analysis["time_deviation"],
            steps_completed=analysis["steps_completed"],
            total_steps=analysis["total_steps"],
            issues_found=analysis["issues_found"],
            critical_issues=analysis["critical_issues"],
            problem_list=analysis["problem_list"],
            recommendations=analysis["recommendations"]
        )
        
        db.add(report)
        db.commit()
        db.refresh(report)
        report_id = report.id
        
        db.close()
        db2 = SessionLocal()
        try:
            return db2.query(DrillReport).get(report_id)
        finally:
            db2.close()
    except Exception as e:
        report_logger.error(f"Failed to save drill report: {str(e)}")
        if 'db' in locals() and db:
            db.rollback()
        return None
    finally:
        if 'db' in locals() and db:
            try:
                db.close()
            except Exception:
                pass


def generate_drill_pdf_report(report: DrillReport, output_path: str) -> Optional[str]:
    db = SessionLocal()
    try:
        report = db.query(DrillReport).options(
            joinedload(DrillReport.drill).joinedload(Drill.function)
        ).get(report.id)
        if not report:
            report_logger.error("DrillReport not found")
            return None
        drill = report.drill

        doc = SimpleDocTemplate(output_path, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=20, alignment=TA_CENTER, spaceAfter=20, textColor=colors.HexColor('#1e40af')
        )
        h2_style = ParagraphStyle(
            'CustomH2', parent=styles['Heading2'],
            fontSize=14, spaceBefore=15, spaceAfter=10, textColor=colors.HexColor('#1e3a8a')
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontSize=10, leading=14
        )
        
        elements = []
        
        elements.append(Paragraph("业务连续性演练报告", title_style))
        elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1e40af')))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("一、演练基本信息", h2_style))
        info_data = [
            ["演练名称", drill.name],
            ["演练类型", drill.drill_type],
            ["开始时间", drill.actual_start_time.strftime("%Y-%m-%d %H:%M:%S") if drill.actual_start_time else "-"],
            ["结束时间", drill.actual_end_time.strftime("%Y-%m-%d %H:%M:%S") if drill.actual_end_time else "-"],
            ["目标恢复时间(分钟)", str(drill.target_recovery_time)],
            ["参与部门", ", ".join(drill.business_units) if drill.business_units else "-"],
        ]
        info_table = Table(info_data, colWidths=[4*cm, 11*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dbeafe')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*cm))
        
        elements.append(Paragraph("二、核心指标", h2_style))
        kpi_data = [
            ["恢复成功率", f"{report.recovery_success_rate}%", "平均恢复时间", f"{report.avg_recovery_time}分钟"],
            ["时间偏差", f"{report.time_deviation}分钟", "步骤完成率", f"{report.steps_completed}/{report.total_steps}"],
            ["发现问题数", str(report.issues_found), "严重问题数", str(report.critical_issues)],
        ]
        kpi_table = Table(kpi_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        kpi_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dcfce7')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#dcfce7')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.3*cm))
        
        if report.problem_list:
            elements.append(Paragraph("三、问题清单", h2_style))
            problem_data = [["序号", "问题类型", "问题描述", "严重程度"]]
            for i, problem in enumerate(report.problem_list, 1):
                problem_data.append([
                    str(i),
                    problem.get("type", "-"),
                    problem.get("description", "-"),
                    problem.get("severity", "-")
                ])
            problem_table = Table(problem_data, colWidths=[1.5*cm, 3*cm, 8.5*cm, 2.5*cm])
            severity_colors = {
                "critical": colors.HexColor('#fecaca'),
                "high": colors.HexColor('#fed7aa'),
                "medium": colors.HexColor('#fef08a'),
                "low": colors.HexColor('#bbf7d0'),
            }
            style_cmds = [
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]
            for i in range(1, len(problem_data)):
                sev = problem_data[i][3]
                if sev in severity_colors:
                    style_cmds.append(('BACKGROUND', (3, i), (3, i), severity_colors[sev]))
            problem_table.setStyle(TableStyle(style_cmds))
            elements.append(problem_table)
            elements.append(Spacer(1, 0.3*cm))
        
        if report.recommendations:
            elements.append(Paragraph("四、改进建议", h2_style))
            for i, rec in enumerate(report.recommendations, 1):
                elements.append(Paragraph(f"{i}. {rec}", normal_style))
                elements.append(Spacer(1, 0.15*cm))
        
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"报告生成时间：{report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}",
                                  ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))
        
        doc.build(elements)
        report_logger.info(f"PDF report generated: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to generate PDF report: {str(e)}", exc_info=True)
        return None
    finally:
        db.close()


def generate_drill_excel_report(report: DrillReport, output_path: str) -> Optional[str]:
    db = SessionLocal()
    try:
        report = db.query(DrillReport).options(
            joinedload(DrillReport.drill).joinedload(Drill.function)
        ).get(report.id)
        if not report:
            report_logger.error("DrillReport not found")
            return None
        drill = report.drill

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        ws1 = wb.active
        ws1.title = "演练概览"
        
        summary_data = [
            ["项目", "内容"],
            ["演练名称", drill.name],
            ["演练类型", drill.drill_type],
            ["开始时间", drill.actual_start_time.strftime("%Y-%m-%d %H:%M:%S") if drill.actual_start_time else "-"],
            ["结束时间", drill.actual_end_time.strftime("%Y-%m-%d %H:%M:%S") if drill.actual_end_time else "-"],
            ["目标恢复时间(分钟)", drill.target_recovery_time],
            ["参与部门", ", ".join(drill.business_units) if drill.business_units else "-"],
            ["恢复成功率", f"{report.recovery_success_rate}%"],
            ["平均恢复时间(分钟)", report.avg_recovery_time],
            ["时间偏差(分钟)", report.time_deviation],
            ["步骤完成数", f"{report.steps_completed}/{report.total_steps}"],
            ["发现问题数", report.issues_found],
            ["严重问题数", report.critical_issues],
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 50
        
        if report.problem_list:
            ws2 = wb.create_sheet("问题清单")
            problem_headers = ["序号", "问题类型", "问题描述", "严重程度"]
            for col_idx, header in enumerate(problem_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            for row_idx, problem in enumerate(report.problem_list, 2):
                ws2.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
                ws2.cell(row=row_idx, column=2, value=problem.get("type", "")).border = thin_border
                ws2.cell(row=row_idx, column=3, value=problem.get("description", "")).border = thin_border
                ws2.cell(row=row_idx, column=4, value=problem.get("severity", "")).border = thin_border
            
            ws2.column_dimensions['A'].width = 8
            ws2.column_dimensions['B'].width = 20
            ws2.column_dimensions['C'].width = 60
            ws2.column_dimensions['D'].width = 15
        
        if report.recommendations:
            ws3 = wb.create_sheet("改进建议")
            ws3.cell(row=1, column=1, value="序号").font = header_font
            ws3.cell(row=1, column=1).fill = header_fill
            ws3.cell(row=1, column=1).border = thin_border
            ws3.cell(row=1, column=2, value="建议内容").font = header_font
            ws3.cell(row=1, column=2).fill = header_fill
            ws3.cell(row=1, column=2).border = thin_border
            
            for row_idx, rec in enumerate(report.recommendations, 2):
                ws3.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
                ws3.cell(row=row_idx, column=2, value=rec).border = thin_border
            
            ws3.column_dimensions['A'].width = 8
            ws3.column_dimensions['B'].width = 80
        
        wb.save(output_path)
        report_logger.info(f"Excel report generated: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to generate Excel report: {str(e)}", exc_info=True)
        return None
    finally:
        db.close()


def generate_drill_reports(drill_id: int) -> Optional[DrillReport]:
    try:
        report = save_drill_report(drill_id)
        if not report:
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = os.path.join(config.REPORT_DIR, f"drill_{drill_id}_report_{timestamp}.pdf")
        excel_path = os.path.join(config.REPORT_DIR, f"drill_{drill_id}_report_{timestamp}.xlsx")
        
        pdf_result = generate_drill_pdf_report(report, pdf_path)
        excel_result = generate_drill_excel_report(report, excel_path)
        
        db = SessionLocal()
        try:
            report.pdf_path = pdf_result
            report.excel_path = excel_result
            db.commit()
        finally:
            db.close()
        
        log_system_event(
            "INFO",
            "report_generation",
            "drill_report_generated",
            {
                "drill_id": drill_id,
                "pdf_path": pdf_result,
                "excel_path": excel_result
            }
        )
        
        return report
    except Exception as e:
        report_logger.error(f"Failed to generate drill reports: {str(e)}", exc_info=True)
        return None


def calculate_monthly_metrics(report_month: str, business_unit: str = None) -> Dict:
    db = SessionLocal()
    try:
        year, month = map(int, report_month.split("-"))
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        query = db.query(Drill).filter(
            Drill.created_at >= start_date,
            Drill.created_at < end_date
        )
        
        if business_unit:
            import json as _json
            all_drills = query.all()
            drills = []
            for d in all_drills:
                if d.business_units and business_unit in d.business_units:
                    drills.append(d)
        else:
            drills = query.all()
        
        completed_drills = [d for d in drills if d.status == "completed"]
        total_drills = len(drills)
        completed_count = len(completed_drills)
        
        drill_completion_rate = (completed_count / total_drills * 100) if total_drills > 0 else 0
        
        recovery_times = []
        for d in completed_drills:
            if d.actual_start_time and d.actual_end_time:
                recovery_times.append((d.actual_end_time - d.actual_start_time).total_seconds() / 60)
        
        avg_recovery_time = (sum(recovery_times) / len(recovery_times)) if recovery_times else 0
        
        wo_query = db.query(ImprovementWorkOrder).filter(
            ImprovementWorkOrder.created_at >= start_date,
            ImprovementWorkOrder.created_at < end_date
        )
        
        if business_unit:
            all_wos = wo_query.all()
            work_orders = [wo for wo in all_wos if wo.assignee_department == business_unit]
        else:
            work_orders = wo_query.all()
        
        total_wos = len(work_orders)
        closed_wos = sum(1 for wo in work_orders if wo.status == "completed")
        improvement_closure_rate = (closed_wos / total_wos * 100) if total_wos > 0 else 0
        
        return {
            "report_month": report_month,
            "business_unit": business_unit,
            "drill_completion_rate": round(drill_completion_rate, 1),
            "avg_recovery_time": round(avg_recovery_time, 1),
            "improvement_closure_rate": round(improvement_closure_rate, 1),
            "total_drills": total_drills,
            "completed_drills": completed_count,
            "total_work_orders": total_wos,
            "closed_work_orders": closed_wos
        }
    except Exception as e:
        report_logger.error(f"Failed to calculate monthly metrics: {str(e)}", exc_info=True)
        return {}
    finally:
        db.close()


def generate_monthly_report(report_month: str) -> Optional[MonthlyReport]:
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        overall_metrics = calculate_monthly_metrics(report_month)
        
        unit_metrics = {}
        for unit in config.BUSINESS_UNITS:
            unit_metrics[unit] = calculate_monthly_metrics(report_month, unit)
        
        trend_data = []
        for i in range(5, -1, -1):
            ref_date = datetime.strptime(report_month, "%Y-%m")
            m = ref_date.month - i
            y = ref_date.year
            while m <= 0:
                m += 12
                y -= 1
            month_str = f"{y}-{m:02d}"
            metrics = calculate_monthly_metrics(month_str)
            metrics["month"] = month_str
            trend_data.append(metrics)
        
        comparison_matrix = {
            "drill_completion_rate": {u: m["drill_completion_rate"] for u, m in unit_metrics.items()},
            "avg_recovery_time": {u: m["avg_recovery_time"] for u, m in unit_metrics.items()},
            "improvement_closure_rate": {u: m["improvement_closure_rate"] for u, m in unit_metrics.items()},
        }
        
        db = SessionLocal()
        try:
            report = MonthlyReport(
                report_month=report_month,
                business_unit=None,
                drill_completion_rate=overall_metrics["drill_completion_rate"],
                avg_recovery_time=overall_metrics["avg_recovery_time"],
                improvement_closure_rate=overall_metrics["improvement_closure_rate"],
                total_drills=overall_metrics["total_drills"],
                completed_drills=overall_metrics["completed_drills"],
                total_work_orders=overall_metrics["total_work_orders"],
                closed_work_orders=overall_metrics["closed_work_orders"],
                trend_data=trend_data,
                comparison_matrix=comparison_matrix
            )
            db.add(report)
            db.commit()
            db.refresh(report)
            
            pdf_path = os.path.join(config.REPORT_DIR, f"monthly_{report_month}_report_{timestamp}.pdf")
            excel_path = os.path.join(config.REPORT_DIR, f"monthly_{report_month}_report_{timestamp}.xlsx")
            
            trend_chart_path = os.path.join(config.REPORT_DIR, f"monthly_{report_month}_trend_{timestamp}.png")
            comp_chart_path = os.path.join(config.REPORT_DIR, f"monthly_{report_month}_comparison_{timestamp}.png")
            
            generate_trend_chart(
                trend_data,
                f"{report_month} 近6个月演练完成率趋势",
                "month",
                "drill_completion_rate",
                trend_chart_path
            )
            
            generate_comparison_chart(
                comparison_matrix["drill_completion_rate"],
                f"{report_month} 各部门演练完成率对比",
                comp_chart_path
            )
            
            pdf_result = generate_monthly_pdf_report(report, pdf_path, trend_chart_path, comp_chart_path)
            excel_result = generate_monthly_excel_report(report, excel_path, unit_metrics)
            
            report.pdf_path = pdf_result
            report.excel_path = excel_result
            db.commit()
            db.refresh(report)
            report_id = report.id
            
            log_system_event(
                "INFO",
                "report_generation",
                "monthly_report_generated",
                {
                    "report_month": report_month,
                    "pdf_path": pdf_result,
                    "excel_path": excel_result
                }
            )
            
            db.close()
            db2 = SessionLocal()
            try:
                return db2.query(MonthlyReport).get(report_id)
            finally:
                db2.close()
        finally:
            if 'db' in locals() and db:
                try:
                    db.close()
                except Exception:
                    pass
    except Exception as e:
        report_logger.error(f"Failed to generate monthly report: {str(e)}", exc_info=True)
        log_system_event("ERROR", "report_generation", "monthly_report_failed", {"error": str(e)}, is_critical=True)
        return None


def generate_monthly_pdf_report(report: MonthlyReport, output_path: str, trend_chart_path: str, comp_chart_path: str) -> Optional[str]:
    try:
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=20, alignment=TA_CENTER, spaceAfter=20, textColor=colors.HexColor('#1e40af')
        )
        h2_style = ParagraphStyle(
            'CustomH2', parent=styles['Heading2'],
            fontSize=14, spaceBefore=15, spaceAfter=10, textColor=colors.HexColor('#1e3a8a')
        )
        
        elements = []
        elements.append(Paragraph(f"{report.report_month} 业务连续性月度报告", title_style))
        elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1e40af')))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("一、总体指标", h2_style))
        kpi_data = [
            ["演练完成率", f"{report.drill_completion_rate}%", "平均恢复时间", f"{report.avg_recovery_time}分钟"],
            ["改进项闭环率", f"{report.improvement_closure_rate}%", "总演练数", str(report.total_drills)],
            ["已完成演练", str(report.completed_drills), "已闭环工单", str(report.closed_work_orders)],
        ]
        kpi_table = Table(kpi_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        kpi_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dcfce7')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#dcfce7')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.3*cm))
        
        if trend_chart_path and os.path.exists(trend_chart_path):
            elements.append(Paragraph("二、趋势分析", h2_style))
            elements.append(Image(trend_chart_path, width=15*cm, height=9*cm))
            elements.append(Spacer(1, 0.3*cm))
        
        elements.append(PageBreak())
        
        if comp_chart_path and os.path.exists(comp_chart_path):
            elements.append(Paragraph("三、部门对比", h2_style))
            elements.append(Image(comp_chart_path, width=15*cm, height=9*cm))
            elements.append(Spacer(1, 0.3*cm))
        
        if report.comparison_matrix:
            elements.append(Paragraph("四、部门对比矩阵", h2_style))
            units = list(report.comparison_matrix["drill_completion_rate"].keys())
            comp_header = ["部门", "演练完成率(%)", "平均恢复时间(分)", "改进项闭环率(%)"]
            comp_data = [comp_header]
            for unit in units:
                comp_data.append([
                    unit,
                    str(report.comparison_matrix["drill_completion_rate"].get(unit, 0)),
                    str(report.comparison_matrix["avg_recovery_time"].get(unit, 0)),
                    str(report.comparison_matrix["improvement_closure_rate"].get(unit, 0))
                ])
            
            comp_table = Table(comp_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
            comp_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(comp_table)
        
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"报告生成时间：{report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}",
                                  ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))
        
        doc.build(elements)
        report_logger.info(f"Monthly PDF report generated: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to generate monthly PDF: {str(e)}", exc_info=True)
        return None


def generate_monthly_excel_report(report: MonthlyReport, output_path: str, unit_metrics: Dict) -> Optional[str]:
    try:
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        ws1 = wb.active
        ws1.title = "总体指标"
        
        summary_data = [
            ["指标", "数值"],
            ["报告月份", report.report_month],
            ["演练完成率", f"{report.drill_completion_rate}%"],
            ["平均恢复时间(分钟)", report.avg_recovery_time],
            ["改进项闭环率", f"{report.improvement_closure_rate}%"],
            ["总演练数", report.total_drills],
            ["已完成演练", report.completed_drills],
            ["总工单数", report.total_work_orders],
            ["已闭环工单", report.closed_work_orders],
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 30
        
        ws2 = wb.create_sheet("部门对比")
        dept_headers = ["部门", "演练完成率(%)", "平均恢复时间(分)", "改进项闭环率(%)", "总演练数", "已完成演练", "总工单数", "已闭环工单"]
        for col_idx, header in enumerate(dept_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        for row_idx, (unit, metrics) in enumerate(unit_metrics.items(), 2):
            values = [
                unit,
                metrics["drill_completion_rate"],
                metrics["avg_recovery_time"],
                metrics["improvement_closure_rate"],
                metrics["total_drills"],
                metrics["completed_drills"],
                metrics["total_work_orders"],
                metrics["closed_work_orders"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
        
        for col in range(1, 9):
            ws2.column_dimensions[chr(64 + col)].width = 18
        
        ws3 = wb.create_sheet("趋势数据")
        if report.trend_data:
            trend_headers = ["月份", "演练完成率(%)", "平均恢复时间(分)", "改进项闭环率(%)", "总演练数", "已完成演练"]
            for col_idx, header in enumerate(trend_headers, 1):
                cell = ws3.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            for row_idx, trend in enumerate(report.trend_data, 2):
                values = [
                    trend.get("month", ""),
                    trend.get("drill_completion_rate", 0),
                    trend.get("avg_recovery_time", 0),
                    trend.get("improvement_closure_rate", 0),
                    trend.get("total_drills", 0),
                    trend.get("completed_drills", 0),
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = center_align
        
        for col in range(1, 7):
            ws3.column_dimensions[chr(64 + col)].width = 20
        
        wb.save(output_path)
        report_logger.info(f"Monthly Excel report generated: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to generate monthly Excel: {str(e)}", exc_info=True)
        return None


def export_lifecycle_data(
    business_unit: str = None,
    drill_type: str = None,
    start_date: datetime = None,
    end_date: datetime = None
) -> str:
    db = SessionLocal()
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(config.EXPORT_DIR, f"lifecycle_export_{timestamp}.xlsx")
        
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        drill_query = db.query(Drill)
        if start_date:
            drill_query = drill_query.filter(Drill.created_at >= start_date)
        if end_date:
            drill_query = drill_query.filter(Drill.created_at < end_date)
        if drill_type:
            drill_query = drill_query.filter(Drill.drill_type == drill_type)
        
        all_drills = drill_query.all()
        
        if business_unit:
            drills = []
            for d in all_drills:
                if d.business_units and business_unit in d.business_units:
                    drills.append(d)
        else:
            drills = all_drills
        
        ws1 = wb.active
        ws1.title = "演练记录"
        drill_headers = ["演练ID", "演练名称", "类型", "状态", "业务单元", "开始时间", "结束时间", "目标恢复时间(分)", "是否模拟"]
        for col_idx, header in enumerate(drill_headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        for row_idx, drill in enumerate(drills, 2):
            values = [
                drill.id,
                drill.name,
                drill.drill_type,
                drill.status,
                ", ".join(drill.business_units) if drill.business_units else "",
                drill.actual_start_time.strftime("%Y-%m-%d %H:%M:%S") if drill.actual_start_time else "",
                drill.actual_end_time.strftime("%Y-%m-%d %H:%M:%S") if drill.actual_end_time else "",
                drill.target_recovery_time,
                "是" if drill.is_simulation else "否"
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
        
        for col in range(1, len(drill_headers) + 1):
            ws1.column_dimensions[chr(64 + col)].width = 18
        
        drill_ids = [d.id for d in drills]
        
        ws2 = wb.create_sheet("演练步骤")
        step_headers = ["演练ID", "步骤号", "描述", "责任方", "目标时长(分)", "状态", "是否超时", "开始时间", "结束时间"]
        for col_idx, header in enumerate(step_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        if drill_ids:
            steps = db.query(DrillStep).filter(DrillStep.drill_id.in_(drill_ids)).all()
            for row_idx, step in enumerate(steps, 2):
                values = [
                    step.drill_id,
                    step.step_number,
                    step.description,
                    step.responsible_party,
                    step.target_duration_minutes,
                    step.status,
                    "是" if step.is_overdue else "否",
                    step.actual_start_time.strftime("%Y-%m-%d %H:%M:%S") if step.actual_start_time else "",
                    step.actual_end_time.strftime("%Y-%m-%d %H:%M:%S") if step.actual_end_time else ""
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = center_align
        
        for col in range(1, len(step_headers) + 1):
            ws2.column_dimensions[chr(64 + col)].width = 18
        
        ws3 = wb.create_sheet("问题记录")
        issue_headers = ["问题ID", "演练ID", "描述", "严重程度", "发现人", "发现时间", "解决方案", "解决时间"]
        for col_idx, header in enumerate(issue_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        if drill_ids:
            issues = db.query(DrillIssue).filter(DrillIssue.drill_id.in_(drill_ids)).all()
            for row_idx, issue in enumerate(issues, 2):
                values = [
                    issue.id,
                    issue.drill_id,
                    issue.description,
                    issue.severity,
                    issue.identified_by or "",
                    issue.identified_at.strftime("%Y-%m-%d %H:%M:%S") if issue.identified_at else "",
                    issue.resolution or "",
                    issue.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if issue.resolved_at else ""
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = center_align
        
        for col in range(1, len(issue_headers) + 1):
            ws3.column_dimensions[chr(64 + col)].width = 20
        
        ws4 = wb.create_sheet("改进工单")
        wo_query = db.query(ImprovementWorkOrder)
        if start_date:
            wo_query = wo_query.filter(ImprovementWorkOrder.created_at >= start_date)
        if end_date:
            wo_query = wo_query.filter(ImprovementWorkOrder.created_at < end_date)
        if business_unit:
            wo_query = wo_query.filter(ImprovementWorkOrder.assignee_department == business_unit)
        
        work_orders = wo_query.all()
        
        wo_headers = ["工单ID", "问题ID", "标题", "描述", "严重程度", "状态", "责任人", "责任部门", "截止日期", "完成日期", "升级级别"]
        for col_idx, header in enumerate(wo_headers, 1):
            cell = ws4.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        for row_idx, wo in enumerate(work_orders, 2):
            values = [
                wo.id,
                wo.issue_id,
                wo.title,
                wo.description or "",
                wo.severity,
                wo.status,
                wo.assignee or "",
                wo.assignee_department or "",
                wo.due_date.strftime("%Y-%m-%d") if wo.due_date else "",
                wo.completed_at.strftime("%Y-%m-%d %H:%M:%S") if wo.completed_at else "",
                wo.escalation_level
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
        
        for col in range(1, len(wo_headers) + 1):
            ws4.column_dimensions[chr(64 + col)].width = 20
        
        wb.save(output_path)
        
        log_system_event(
            "INFO",
            "report_generation",
            "lifecycle_data_exported",
            {
                "output_path": output_path,
                "business_unit": business_unit,
                "drill_type": drill_type,
                "drills_count": len(drills),
                "work_orders_count": len(work_orders)
            }
        )
        
        report_logger.info(f"Lifecycle data exported: {output_path}")
        return output_path
    except Exception as e:
        report_logger.error(f"Failed to export lifecycle data: {str(e)}", exc_info=True)
        return ""
    finally:
        db.close()
